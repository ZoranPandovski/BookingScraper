class FileWriter:
    #TODO: cleanup output_file
    def __init__(self, data, out_format=None, country='Macedonia'):
        self.data = data
        self.format = out_format
        self.country= country

    def output_file(self):
        '''
        Write list of hotels in file
        :return:
        '''
        format = self.format.lower()
        file_name = ''
        if format == 'json' or format is None:
            import json
            file_name = 'hotels-in-{country}.txt'.format(
                country=self.country.replace(" ", "-"))
            with open(file_name, 'w', encoding='utf-8') as outfile:
                json.dump(list(self.data), outfile, indent=2, ensure_ascii=False)
        elif format == 'excel':
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active

            heading1 = '#'
            heading2 = 'Accommodation'
            ws.cell(row=1, column=1).value = heading1
            ws.cell(row=1, column=2).value = heading2

            for i, item in enumerate(self.data):
                # Extract number and title from string
                tokens = item.split()
                n = tokens[0]
                title = ' '.join(tokens[2:])

                ws.cell(row=i + 2, column=1).value = n
                ws.cell(row=i + 2, column=2).value = title

            file_name = 'hotels-in-{country}.xls'.format(
                country=self.country.replace(" ", "-"))
            wb.save(file_name)
        elif format == 'csv':
            file_name = 'hotels-in-{country}.csv'.format(
                country=self.country.replace(" ", "-"))
            with open(file_name, 'w', encoding='utf-8') as outfile:
                for i, item in enumerate(self.data):
                    # Extract number and title from string
                    tokens = item.split()
                    n = tokens[0]
                    title = ' '.join(tokens[2:])

                    s = n + ', ' + title + '\n'
                    outfile.write(s)

        return file_name